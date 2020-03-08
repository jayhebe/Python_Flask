import openpyxl


def read_head(excel_file):
    excel_head = {}

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    for col in range(1, ws.max_column + 1):
        excel_head[col] = ws.cell(row=1, column=col).value

    return excel_head


def read_content(excel_file, list_columns):
    result = []

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    for col in list_columns:
        col_result = []
        for row in range(1, ws.max_row + 1):
            col_result.append(ws.cell(row=row, column=col).value)

        result.append(col_result)

    result = [[row[i] for row in result] for i in range(len(result[0]))]

    return result
